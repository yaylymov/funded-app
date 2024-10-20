from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from openpyxl import load_workbook
from fastapi.middleware.cors import CORSMiddleware
import logging

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

file_path = 'data/funded-database.xlsx'
wb = load_workbook(file_path)
ws = wb.active


class Questionnaire(BaseModel):
    state: str
    companySize: str
    areas: str
    grantsAmount: int
    revenue: int


def filter_grants(answers):
    """
    This function filters grants from the Excel sheet based on the given answers.
    """
    filtered_grants = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(cell is None for cell in row):
            # Skip rows with missing data
            continue

        grant_option, state, company_size, area, grant_volume, funding_quota, approval_rate, benefit_cost_score = row

        # For debugging reasons
        print(
            f"Comparing: state='{answers.state}' with '{state}', companySize='{answers.company_size}' with '{company_size}', areas='{answers.areas_active}' with '{area}'")

        # Filter logic based on the answers
        if (answers.state.lower() in state.lower()) and \
                (answers.company_size.lower() in company_size.lower()) and \
                (answers.areas_active.lower() in area.lower()):
            filtered_grants.append({
                "funding_option": grant_option,
                "grant_volume": grant_volume,
                "funding_quota": funding_quota,
                "approval_rate": approval_rate,
                "benefit_cost_score": benefit_cost_score
            })

    return filtered_grants


@app.post("/questions")
async def process_questionnaire(data: Questionnaire):
    logger.info(f"Received data: {data}")
    try:
        filtered_grants = filter_grants(data)
        logger.info(f"Filtered grants: {filtered_grants}")
        return filtered_grants
    except Exception as e:
        logger.error(f"An error occurred: {str(e)}")
        raise HTTPException(status_code=500, detail=f"An error occurred: {str(e)}")


@app.get("/")
async def root():
    return {"message": "FUNDED backend is up and running!"}
